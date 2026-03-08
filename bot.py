import pandas as pd
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

# ===============================
# حط هنا الـ TOKEN بتاعك
# ===============================
TOKEN = "8552659968:AAHx_lbZwxvOmPfWGuhhUMlJmIDIsuNlmw0"

file_path = None  # هيتخزن فيه الملف المرسل

# ===============================
# Start Command
# ===============================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Welcome! Send me your Excel file and then type /run to process it.")

# ===============================
# استقبال الملف
# ===============================
async def receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global file_path
    file = await update.message.document.get_file()
    file_path = "input.xlsx"
    await file.download_to_drive(file_path)
    await update.message.reply_text("File received. Now type /run to process it.")

# ===============================
# تشغيل الكود على الملف
# ===============================
async def run_code(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global file_path
    if not file_path:
        await update.message.reply_text("Please send a file first!")
        return

    # ===============================
    # اقرأ Excel
    # ===============================
    df = pd.read_excel(file_path, sheet_name=0)

    # ===============================
    # دالة التصنيف
    # ===============================
    def classify_description(row):
        desc_col = 'Transaction Remarks'
        type_col = 'Transaction Code Description'
        refer_col ='Transaction Reference'
        
        desc = str(row.get(desc_col, '')).upper().strip() if pd.notna(row.get(desc_col)) else ""
        refer = str(row.get(refer_col, '')).upper().strip() if pd.notna(row.get(refer_col)) else ""
        trans_type = str(row.get(type_col, '')).upper().strip() if pd.notna(row.get(type_col)) else ""
        account = str(row.get('Account', '')).strip()
    
        if 'AMEX (MIDDLE EAST) BSC' in desc:
            return "AMEX"
        if 'BOTIM MONEY TECHNOLOGY LLC' in desc or 'PAYBY TECHNOLOGY PROJECTS LLC' in desc:
            return "Astra"
        if 'INWARD REMITTANCE CHARGE' in desc:
            return "Bank Charges"
        if ('Account Number:0353418581002' in account and 'CHARGE COLLECTION - INCL VAT' in desc ) \
            or ('Account Number:0353418581001' in account and 'CHARGE COLLECTION - INCL VAT' in desc) \
            or ('Account Number:0353418581001' in account and 'CHARGE COLLECTION-INCL' in desc):
            return "ACH Fees"
        if 'CHECKOUT MENA FZ-LLC' in desc:
            return "Checkout"
        if ('Account Number:0353418581002' in account and ('FUNDS TRANSFER WITHIN RAKBANK' in desc or 'AANI TO' in desc)):
            return "Instant settlement"
        if 'FUNDS TRANSFER BETWEEN OWN ACCOUNTS' in desc:
            return "Intercompany"
        if 'MASHREQBANK' in desc:
            return "Mashriq Merchant Sett."
        if '"MERCHANTEMPORARY MERCHANT PAYMENT"' in desc:
            return "Mashriq Merchant Sett. USD"
        if ('RAKBANK MERCH STLMT' in desc and ('3777' in desc or '4426' in desc or '4427' in desc or '4435' in desc or '4422' in desc or '4438' in desc or '4432' in desc or '4436' in desc or '4419' in desc or '3756' in desc or '4418' in desc or '4408' in desc or '5632' in desc or '5707' in desc ) ):
            return "Rak Pos"
        if ('RAKBANK MERCH STLMT' in desc and ('4434' in desc or '4437' in desc or '4439' in desc or '4425' in desc or '4428' in desc or '4431' in desc or '4423' in desc or '3550' in desc or '0472' in desc or '0473' in desc or '0497' in desc or '1319' in desc or '0495' in desc or '0499' in desc or '0501' in desc or '1318' in desc or '2219' in desc or '2448' in desc or '3757' in desc  or '3328' in desc or '3326' in desc or '3333' in desc  ) ):
            return "Rak Ecom"
        if 'TABBY' in desc:
            return "Tabby"
        if 'TAMARA' in desc:
            return "Tamara"
        if 'FUNDS TRANSFER WITHIN RAKBANK' in desc or 'OUTWARD T/T' in desc:
            return "UAE Merchant Settl."
        return "Unclassified"

    # ===============================
    # طبق التصنيف
    # ===============================
    df['Class'] = df.apply(classify_description, axis=1)

    # ===============================
    # تحويل التاريخ
    # ===============================
    if pd.api.types.is_numeric_dtype(df['Date']):
        df['Date'] = pd.to_datetime(df['Date'], origin='1899-12-30', unit='D')
    else:
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')

    # ===============================
    # Summary Column
    # ===============================
    df['Summary_Amount'] = df.apply(lambda r: r['Withdrawals'] if r['Class']=="UAE Merchant Settl." else r['Deposits'], axis=1)

    # ===============================
    # Pivot Summary
    # ===============================
    summary = df.groupby(['Date', 'Class'])['Summary_Amount'].sum().reset_index()
    pivot_summary = summary.pivot(index='Date', columns='Class', values='Summary_Amount').fillna(0)

    start_date = df['Date'].min().replace(day=1)
    end_date = df['Date'].max().replace(day=1) + pd.offsets.MonthEnd(0)
    all_dates = pd.date_range(start=start_date, end=end_date, freq='D')
    pivot_summary = pivot_summary.reindex(all_dates, fill_value=0)
    pivot_summary.index.name = 'Date'

    # ترتيب الأعمدة
    desired_order = ["Rak Pos","Rak Ecom", "Checkout", "AMEX", "Astra", "Tabby", "Tamara","UAE Merchant Settl."]
    for col in desired_order:
        if col not in pivot_summary.columns:
            pivot_summary[col] = 0
    pivot_summary = pivot_summary[desired_order]

    # ===============================
    # حفظ Output
    # ===============================
    output_file = file_path.replace(".xlsx", "_classified.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        pivot_summary.to_excel(writer, sheet_name='Summary')

    # ===============================
    # تلوين العنوان
    # ===============================
    wb = load_workbook(output_file)
    ws = wb['Summary']
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    wb.save(output_file)

    # ===============================
    # ارسال الملف للمستخدم
    # ===============================
    await update.message.reply_document(document=open(output_file, "rb"))

    print(f"DONE: {output_file}")
    print(df['Class'].value_counts())

# ===============================
# بناء البوت
# ===============================
app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("run", run_code))
app.add_handler(MessageHandler(filters.Document.ALL, receive_file))
app.run_polling()